#!/usr/bin/env python

from openpyxl import load_workbook
from pprint import pprint
import json
import os

CUR_DIR = os.path.dirname(__file__)

BALANCE_EXCEL = os.path.join(
            CUR_DIR,
            'balance.xlsm')

BALANCE_JSON = os.path.join(
            CUR_DIR,
            '..', 'back', 'balance',
            'balance.json')

PRIVATE_BALANCE_JSON = os.path.join(
            CUR_DIR,
            '..', 'back', 'balance',
            'private_balance.json')

LOCAL_BALANCE_JSON = os.path.join(
            CUR_DIR,
            'balance.json')

LOCAL_PRIVATE_BALANCE_JSON = os.path.join(
            CUR_DIR,
            'private_balance.json')

STRATEGIES_FOLDER = os.path.join(CUR_DIR, 'strategies')
BASES_FOLDER = os.path.join(CUR_DIR, 'bases')
BALANCE_BASES_FOLDER = os.path.join(
            CUR_DIR,
            '..', 'back', 'balance',
            'bases')

EXTRAS_TO_PROPS = (
    ('extra.cr.store.{}', 'crystaliteCapacity'),
    ('extra.cr.farm.{}', 'crystaliteProduction'),
    ('extra.cr.max.{}', 'crystaliteProductionCapacity'),

    ('extra.ad.store.{}', 'adamantiteCapacity'),
    ('extra.ad.farm.{}', 'adamantiteProduction'),
    ('extra.ad.max.{}', 'adamantiteProductionCapacity'),

    ('extra.ti.store.{}', 'titaniumCapacity'),
    ('extra.ti.farm.{}', 'titaniumProduction'),
    ('extra.ti.max.{}', 'titaniumProductionCapacity'),

    ('extra.gen.en.{}', 'generateEnergy'),

)

BOTS_PROPS = (
    ('bt.building', 'building'),
    ('bt.cr.farm', 'crystaliteProduction'),
    ('bt.ad.farm', 'adamantiteProduction'),
    ('bt.ti.farm', 'titaniumProduction'),
    ('bt.research', 'research'),
    ('bt.dec.time', 'decExpeditionTime'),
)

CHEST_LEVELS = ('lvl1', 'lvl2', 'lvl3')

wb = load_workbook(BALANCE_EXCEL, data_only=True)

ROW_TYPE = 1
ROW_NAMES = 2
ROW_LEVEL = (3, 17)
SHEETS_WITH_NUMBERS = ('Numbers', 'Missions', 'Bots', 'Battle')
data = {}

private_config = {
    'garbage': {

    },
    'chest': {
        'lvl1': {
            'order': ['crystalite', 'adamantite', 'titanium', 'coin', 'bot', 'infantryBot', 'rocketBot', 'heavyBot', 'mine', 'rocket', 'heal', 'power', 'm.common', 'm.rare', 'm.epic', 'm.legend']
        },
        'lvl2': {
            'order': ['crystalite', 'adamantite', 'titanium', 'coin', 'bot', 'infantryBot', 'rocketBot', 'heavyBot', 'mine', 'rocket', 'heal', 'power', 'm.common', 'm.rare', 'm.epic', 'm.legend']
        },
        'lvl3': {
            'order': ['crystalite', 'adamantite', 'titanium', 'coin', 'bot', 'infantryBot', 'rocketBot', 'heavyBot', 'mine', 'rocket', 'heal', 'power', 'm.common', 'm.rare', 'm.epic', 'm.legend']
        }
    },
    'coded_chest': {
        'lvl1': [],
        'lvl2': [],
        'lvl3': []
    }
}

result = {
    'commandCenter':{
        'size': {'x': 4, 'y': 4},
        'role': 'center'
    },
    'crystaliteSilo':{
        'size': {'x': 3, 'y': 3},
        'category': 'resource',
        'position': 1,
        'role': 'building'
    },
    'crystaliteFarm':{
        'size': {'x': 3, 'y': 3},
        'category': 'resource',
        'position': 2,
        'role': 'building'
    },
    'adamantiteStorage':{
        'size': {'x': 3, 'y': 3},
        'category': 'resource',
        'position': 3,
        'role': 'building'
    },
    'adamantiteMine':{
        'size': {'x': 3, 'y': 3},
        'category': 'resource',
        'position': 4,
        'role': 'building'
    },
    'titaniumStorage':{
        'size': {'x': 3, 'y': 3},
        'category': 'resource',
        'position': 6,
        'role': 'building'
    },
    'titaniumLab':{
        'size': {'x': 3, 'y': 3},
        'category': 'resource',
        'position': 7,
        'role': 'building'
    },
    'vault':{
        'size': {'x': 3, 'y': 3},
        'category': 'resource',
        'position': 5,
        'role': 'building'
    },
    'electronic':{
        'size': {'x': 2, 'y': 2},
        'category': 'resource',
        'position': 5,
        'role': 'building'
    },
    # 'collector':{
    #     'size': {'x': 3, 'y': 3},
    #     'category': 'resource',
    #     'position': 8
    # },
    'craftPad':{
        'size': {'x': 2, 'y': 2},
        'category': 'support',
        'position': 1,
        'role': 'building'
    },
    'laboratory':{
        'size': {'x': 3, 'y': 3},
        'category': 'support',
        'position': 6,
        'role': 'building'
    },
    'bots':{
        'size': {'x': 2, 'y': 2},
        'category': 'support',
        'position': 2,
        'role': 'building'
    },
    'radar':{
        'size': {'x': 3, 'y': 3},
        'category': 'support',
        'position': 4,
        'role': 'building'
    },
    'garbage':{
        'size': {'x': 3, 'y': 3},
        'category': 'support',
        'position': 5,
        'role': 'building'
    },
    'flagman':{
        'size': {'x': 2, 'y': 2},
        'category': 'support',
        'position': 7,
        'role': 'building'
    },
    'sentryGun':{
        'size': {'x': 3, 'y': 3},
        'category': 'defense',
        'position': 1,
        'role': 'tower'
    },
    'machineGun':{
        'size': {'x': 2, 'y': 2},
        'category': 'defense',
        'position': 2,
        'role': 'tower'
    },
    'mine':{
        'size': {'x': 1, 'y': 1},
        'category': 'defense',
        'position': 3
    },
    'defPlatform':{
        'size': {'x': 2, 'y': 2},
        'category': 'defense',
        'position': 4,
        'role': 'tower'
    },
    'rocketGun':{
        'size': {'x': 2, 'y': 2},
        'category': 'defense',
        'position': 5,
        'role': 'tower'
    },
    'weather':{
        'size': {'x': 3, 'y': 3},
        'role': 'building'
    },
    'goldenCat':{'size': {'x': 2, 'y': 2},
        'counts': [{
            'count': 1,
            'commandCenter': 1
        }],
        'stats': [{'display': {'hitpoints': 100},
                    'level': 1,
                    'requiredPlayerLevel': 1,
                    'upgradeCost': {'coin': 5000},
                    'upgradeTime': 0,
                    'xpGain': 0}],
        'category': 'decor',
        'position': 1
    },
    'rocketStone':{'size': {'x': 2, 'y': 2},
        'counts': [{
            'count': 1,
            'commandCenter': 1
        }],
        'stats': [{'display': {'hitpoints': 100},
                    'level': 1,
                    'requiredPlayerLevel': 1,
                    'upgradeCost': {'adamantite': 10000},
                    'upgradeTime': 0,
                    'xpGain': 0}],
        'category': 'decor',
        'position': 2
    },
    'scarecrow':{'size': {'x': 2, 'y': 2},
        'counts': [{
            'count': 1,
            'commandCenter': 1
        }],
        'stats': [{'display': {'hitpoints': 100},
                    'level': 1,
                    'requiredPlayerLevel': 1,
                    'upgradeCost': {'adamantite': 150000},
                    'upgradeTime': 0,
                    'xpGain': 0}],
        'category': 'decor',
        'position': 3
    },
    'ufo':{'size': {'x': 3, 'y': 3},
        'counts': [{
            'count': 1,
            'commandCenter': 1
        }],
        'stats': [{'display': {'hitpoints': 100},
                    'level': 1,
                    'requiredPlayerLevel': 1,
                    'upgradeCost': {'coin': 250},
                    'upgradeTime': 0,
                    'xpGain': 0}],
        'category': 'decor',
        'position': 4
    },
    'zeus':{'size': {'x': 3, 'y': 3},
        'counts': [{
            'count': 1,
            'commandCenter': 1
        }],
        'stats': [{'display': {'hitpoints': 100},
                    'level': 1,
                    'requiredPlayerLevel': 1,
                    'upgradeCost': {'coin': 1000},
                    'upgradeTime': 0,
                    'xpGain': 0}],
        'category': 'decor',
        'position': 5
    },
    'dragonBones':{'size': {'x': 2, 'y': 2},
        'counts': [{
            'count': 1,
            'commandCenter': 1
        }],
        'stats': [{'display': {'hitpoints': 100},
                    'level': 1,
                    'requiredPlayerLevel': 1,
                    'upgradeCost': {'adamantite': 250000},
                    'upgradeTime': 0,
                    'xpGain': 0}],
        'category': 'decor',
        'position': 6
    },
    'flagStock':{'size': {'x': 2, 'y': 2},
        'counts': [{
            'count': 1,
            'commandCenter': 1
        }],
        'role': 'building',
        'stats': [{'display': {'hitpoints': 100},
                    'level': 1,
                    'requiredPlayerLevel': 1,
                    
                    'upgradeCost': {},
                    'upgradeTime': 0,
                    'xpGain': 0}]
    },
    'obstacle2':{
        'size': {'x': 2, 'y': 2},
        'category': 'obstacle',
        'role': 'obstacle'
    },
}



game_config = {
    'buildings': result,
    'initial_codes': {
        'python-3': {

            'craft.py': '''
from battle import commander


craft_client = commander.CraftClient()
craft_client.do_land_units()

def unit_landed(data):
    unit_client = commander.UnitClient(data['id'])

    def search_and_destroy(data=None):
        enemy = unit_client.ask_nearest_enemy()
        unit_client.do_attack(enemy['id'])
        unit_client.when_im_idle(search_and_destroy)

    search_and_destroy()

craft_client.when_unit_landed(unit_landed)            
'''.strip(),

            'tower.py': '''
from battle import commander


tower_client = commander.Client()

def search_next_target(data, **kwargs):
    enemies = tower_client.ask_enemy_items_in_my_firing_range()
    if enemies:
        unit_in_firing_range(enemies[0])
    else:
        tower_client.when_enemy_in_range(unit_in_firing_range)

def unit_in_firing_range(data, **kwargs):
    tower_client.attack_item(data['id'])
    tower_client.when_im_idle(search_next_target)

tower_client.when_enemy_in_range(unit_in_firing_range)
'''.strip()
        },
        'js-node': {

        }
    }
}

def scan_data(result, data, sheets, levels):

    def fill_levels(ws, type_name, col_start, col_end):
        if result and type_name not in result:
            return
        if type_name not in data:
            data[type_name] = []

        properties = []
        for cell in ws[col_start+str(ROW_NAMES) : col_end+str(ROW_NAMES)][0]:
            if cell.value:
                properties.append(cell.value)
        for lvl_id, lvl_row in enumerate(range(levels[0], levels[1] + 1)):
            try:
                data_level = data[type_name][lvl_id]
            except IndexError:
                data_level = {
                    'center': lvl_id + 1
                }
                data[type_name].append(data_level)
            for prop, cell in zip(properties, ws[col_start+str(lvl_row) : col_end+str(lvl_row)][0]):
                data_level[prop.lower().replace(' ', '')] = cell.value
        

    for ws_name in sheets:
        ws = wb[ws_name]
        type_name = None
        col_start = None
        col_end = None

        for cell in ws[str(ROW_TYPE):str(ROW_TYPE)]:
            if cell.value:
                if type_name:
                    fill_levels(ws, type_name, col_start, col_end)
                type_name = cell.value
                col_end = col_start = cell.column
            else:
                col_end = cell.column

        fill_levels(ws, type_name, col_start, col_end)

scan_data(result, data, SHEETS_WITH_NUMBERS, ROW_LEVEL)

#pprint(data)

for b_type, b_data in data.items():
    r_data = result[b_type]
    r_counts = r_data['counts'] = []
    r_stats = r_data['stats'] = []
    for lvl_data in b_data:
        if lvl_data.get('inc'):
            r_counts.append({
                'count': lvl_data['total'],
                'commandCenter': lvl_data['center']
                })
        if lvl_data.get('has'):
            new_stat = {
                'level': lvl_data['level'],
                'xpGain': lvl_data.get('xpgain'),
                'upgradeTime': lvl_data['uptime'],
            }
            if b_type != 'commandCenter':
                new_stat['commandCenter'] = lvl_data['center'] 
            r_stats.append(new_stat)
            if 'lvlreq' in lvl_data:
                new_stat['requiredPlayerLevel'] = lvl_data['lvlreq']

            if 'energy' in lvl_data:
                new_stat['energy'] = lvl_data['energy']

            if lvl_data.get('bt.size'):
                new_stat['botMax'] = lvl_data['bt.size']
                bot_update = new_stat['botImprove'] = {}
                for data_key, stat_key in BOTS_PROPS:
                    if lvl_data.get(data_key):
                        bot_update[stat_key] = lvl_data[data_key]

            if 'bt.space' in lvl_data:
                new_stat['botSpace'] = lvl_data['bt.space']


            cost = new_stat['upgradeCost'] = {}
            if lvl_data.get('ad.price'):
                cost['adamantite'] = lvl_data['ad.price']
            
            if lvl_data.get('cr.price'):
                cost['crystalite'] = lvl_data['cr.price']

            if lvl_data.get('ti.price'):
                cost['titanium'] = lvl_data['ti.price']
            
            if 'coof.time' in lvl_data:
                new_stat['coin'] = {
                    'time': lvl_data['coof.time'] * 1000,
                    'adamantite': lvl_data['coof.ad.res'] * 1000,
                    'crystalite': lvl_data['coof.cr.res'] * 1000,
                    'titanium': lvl_data['coof.ti.res'] * 1000
                }
            missions = []
            for i in range(1, 6):
                mission_num = 'mission.{}'.format(i)
                if lvl_data.get(mission_num):
                    mission = {
                        'num': i,
                        'mission': lvl_data[mission_num],
                        'codeXp': lvl_data['cxp.{}'.format(i)],
                        'codeLevel': lvl_data['lreq.{}'.format(i)],
                        'display': {}
                    }
                    missions.append(mission)

                    for xls_attr_tmp, display_prop in EXTRAS_TO_PROPS:
                        xls_attr = xls_attr_tmp.format(i)
                        if not lvl_data.get(xls_attr):
                            continue
                        mission['display'][display_prop] = lvl_data[xls_attr]
            
            if missions:
                new_stat['missions'] = missions

            ret = {}
            if 'ad.return' in lvl_data:
                ret['adamantite'] = lvl_data['ad.return']
            
            if ret:
                new_stat['ret'] = ret

            display = new_stat['display'] = {}

            if 'cr.farm' in lvl_data:
                display['crystaliteProduction'] = lvl_data['cr.farm']
                display['crystaliteProductionCapacity'] = lvl_data['cr.max']
            if 'ad.farm' in lvl_data:
                display['adamantiteProduction'] = lvl_data['ad.farm']
                display['adamantiteProductionCapacity'] = lvl_data['ad.max']
            if 'ti.farm' in lvl_data:
                display['titaniumProduction'] = lvl_data['ti.farm']
                display['titaniumProductionCapacity'] = lvl_data['ti.max']
            if 'cr.store' in lvl_data:
                display['crystaliteCapacity'] = lvl_data['cr.store']
            if 'ad.store' in lvl_data:
                display['adamantiteCapacity'] = lvl_data['ad.store']
            if 'ti.store' in lvl_data:
                display['titaniumCapacity'] = lvl_data['ti.store']
            if 'gen.en' in lvl_data:
                display['generateEnergy'] = lvl_data['gen.en']

            if 'chargingtime' in lvl_data:
                display['charging_time'] = lvl_data['chargingtime']
            if 'hitpoints' in lvl_data:
                display['hit_points'] = lvl_data['hitpoints']
            if 'damagepershot' in lvl_data:
                display['damage_per_shot'] = lvl_data['damagepershot']
            if 'firerange' in lvl_data:
                display['firing_range'] = lvl_data['firerange']
            if 'firerange100' in lvl_data:
                display['firing_range_100'] = lvl_data['firerange100']
            if 'startchance' in lvl_data:
                display['start_chance'] = lvl_data['startchance']
            if 'rateoffire' in lvl_data:
                display['rate_of_fire'] = lvl_data['rateoffire']
            if 'chargesize' in lvl_data:
                display['charge_size'] = lvl_data['chargesize']
            if 'a_rocket' in lvl_data:
                display['has_rocket'] = lvl_data['a_rocket']
                display['has_heal'] = lvl_data['a_heal']
                display['has_power'] = lvl_data['a_power']

            if 'fieldofview' in lvl_data:
                display['field_of_view'] = lvl_data['fieldofview']
            if 'rateofturn' in lvl_data:
                display['rate_of_turn'] = lvl_data['rateofturn']
            if 'damagepersecond' in lvl_data:
                display['damage_per_second'] = lvl_data['damagepersecond']
            if 'firingtimelimit' in lvl_data:
                display['firing_time_limit'] = lvl_data['firingtimelimit']
            if 'fullcooldowntime' in lvl_data:
                display['full_cooldown_time'] = lvl_data['fullcooldowntime']
            if 'minpercentageafteroverheat' in lvl_data:
                display['min_percentage_after_overheat'] = lvl_data['minpercentageafteroverheat']
            if 'rocketspeed' in lvl_data:
                display['rocket_speed'] = lvl_data['rocketspeed']
            if 'rocketexplosionradius' in lvl_data:
                display['rocket_explosion_radius'] = lvl_data['rocketexplosionradius']
            
            if 'slots' in lvl_data:
                display['slots'] = lvl_data['slots']

            if 'unitcapacity' in lvl_data:
                display['unitCapacity'] = lvl_data['unitcapacity']

            if b_type == 'laboratory':
                new_stat['research'] = {
                    'mine': lvl_data['max.mine'],
                    'infantryBot': lvl_data['max.infantrybot'],
                    'rocketBot': lvl_data['max.rocketbot'],
                    'heavyBot': lvl_data['max.heavybot'],
                    'rocket': lvl_data['max.a_rocket'],
                    'heal': lvl_data['max.a_heal'],
                    'power': lvl_data['max.a_power'],
                }
            
            if b_type in ('craftPad', 'flagman'):
                display['expSpeed'] = round(lvl_data['speed'], 1)

            if b_type == 'garbage':
                display['expSpeed'] = round(lvl_data['speed'], 1)
                new_stat['expPrice'] = {
                    'crystalite': lvl_data['ex.cr.price']
                }
                for term in ('st.lvl1', 'st.lvl2', 'st.lvl3', 
                             'lt.lvl1', 'lt.lvl2', 'lt.lvl3'):
                    
                    display[term.replace('st', 'short').replace('lt', 'long')] = int(lvl_data[term]*100)
                new_stat['shortExpedition'] = {
                    'time': lvl_data['shorttime'],
                    'price': {
                        'crystalite': lvl_data['shortprice']   
                    }
                }
                new_stat['longExpedition'] = {
                    'time': lvl_data['longtime'],
                    'price': {
                        'crystalite': lvl_data['longprice']   
                    }
                }

                private_config['garbage'][lvl_data['level']] = {
                    'shortTimeCodingPos': int(lvl_data['st.codingpos']*100),
                    'longTimeCodingPos': int(lvl_data['lt.codingpos']*100)
                }

from copy import deepcopy
result['obstacle3'] =  deepcopy(result['obstacle2'])
result['obstacle3']['size'] = {
    'x': 3,
    'y': 3,
}
result['obstacle5'] =  deepcopy(result['obstacle2'])
result['obstacle5']['size'] = {
    'x': 5,
    'y': 5,
}

def collect_table(ws_head, ws_data):
    names = {}
    for cell in ws_head[0]:
        names[cell.column] = cell.value

    for row in ws_data:
        if not row[0].value:
            break
        level = {}
        for cell in row:
            if cell.value is not None:
                level[names[cell.column]] = cell.value

        yield level

def set_raw_levels(data):
    for item in data:
        item.pop('coof', None)
        item.pop('grow', None)
        yield {
            'level': item.pop('level'),
            'exp': round(item.pop('exp')),
            'rewards': item
        }

ws = wb['Levels']
game_config['levels'] = [{
    'level': 0,
    'exp': 0,
    'rewards': {}
}] + list(set_raw_levels(collect_table(ws['A2': 'G2'], ws['A3':'G100'])))

ws = wb['codeLevels']
game_config['codeLevels'] = [{
    'level': 0,
    'exp': 0,
    'rewards': {}
}] + list(set_raw_levels(collect_table(ws['A2': 'E2'], ws['A3':'E100'])))


labs = {
    'infantryBot': {},
    'rocketBot': {},
    'heavyBot': {},
    'mine': {}
}
game_config['units'] = labs

labs_data = {}
scan_data(labs, labs_data, ('Units',), (3, 29))

for name, unit in labs.items():
    stats = []
    unit['stats'] = stats
    for raw in labs_data[name]:
        if not raw['level']:
            continue
        raw.pop('center')
        raw['price'] = {
            'crystalite': raw.pop('crystalite')
        }
        raw['upTime'] = raw.pop('uptime')

        raw['hit_points'] = raw.pop('hitpoints', 0)
        raw['damage_per_shot'] = raw.pop('damagepershot', 0)
        raw['firing_range'] = raw.pop('firerange', 0)
        if 'rateoffire' in raw:
            raw['rate_of_fire'] = raw.pop('rateoffire')
        if 'chargingtime' in raw:
            raw['charging_time'] = raw.pop('chargingtime')

        if 'rateofturn' in raw:
            raw['rate_of_turn'] = raw.pop('rateofturn')
        if 'damagepersecond' in raw:
            raw['damage_per_second'] = raw.pop('damagepersecond')
        if 'firingtimelimit' in raw:
            raw['firing_time_limit'] = raw.pop('firingtimelimit')
        if 'fullcooldowntime' in raw:
            raw['full_cooldown_time'] = raw.pop('fullcooldowntime')
        if 'minpercentageafteroverheat' in raw:
            raw['min_percentage_after_overheat'] = raw.pop('minpercentageafteroverheat')
        if 'rocketspeed' in raw:
            raw['rocket_speed'] = raw.pop('rocketspeed')

        if 'hireprice' in raw:
            raw['hirePrice'] = {
                'crystalite': raw.pop('hireprice')
            }
            raw['hireTime'] = raw.pop('hiretime')
        raw.pop('k', None)
        stats.append(raw)

flag = {
    'rocket': {},
    'heal': {},
    'power': {}
}

flag_data = {}
scan_data(flag, flag_data, ('Flagman',), (3, 16))

for name, unit in flag.items():
    stats = []
    unit['stats'] = stats
    for raw in flag_data[name]:
        if not raw['level']:
            continue
        raw['price'] = {
            'crystalite': raw.pop('crystalite')
        }
        raw['upTime'] = raw.pop('uptime')
        stats.append(raw)

labs.update(flag)


ws = wb['Quests']
game_config['quests'] = {}

for pos, row in enumerate(collect_table(ws['A1': 'M1'], ws['A2':'M100'])):
    quest = {}
    name = row.pop('name')
    for par in ('parameter', 'amount', 'translate', 'lvl'):
        quest[par] = row.pop(par)
    quest['parent'] = row.pop('parent', None)
    quest['rewards'] = row
    quest['pos'] = pos

    game_config['quests'][name] = quest

#pprint(result)

private_data = {}
scan_data(None, private_data, ('Chests',), ROW_LEVEL)

PROPERTIES = (
    ('cr', 'crystalite'),
    ('ad', 'adamantite'),
    ('ti', 'titanium'),
    ('coin', 'coin'),
    ('bot', 'bot'),
    ('infantrybot', 'infantryBot'),
    ('rocketbot', 'rocketBot'),
    ('heavybot', 'heavyBot'),
    ('mine', 'mine'),
    ('rocket', 'rocket'),
    ('heal', 'heal'),
    ('power', 'power'),
    ('m.common', 'm.common'),
    ('m.rare', 'm.rare'),
    ('m.epic', 'm.epic'),
    ('m.legend', 'm.legend'),
)

for chest in CHEST_LEVELS:
    chest_data = private_data[chest]
    chest_confing = private_config['chest'][chest]['cc'] = {}
    for lvl, chest_data in enumerate(private_data[chest], 1):
        lvl_config = chest_confing[lvl] = {}
        for prop_data, prop_conf in PROPERTIES:
            pos = chest_data[prop_data + '.pos'] * 100
            lvl_config[prop_conf] = {
                'posibility': pos,
                'min': pos and chest_data.get(prop_data + '.min', 1) or 0,
                'max': pos and chest_data.get(prop_data + '.max', 1) or 0
            }


ws = wb['CodedChest']

for row in ws['A2': 'C100']:
    for lvl, cell in zip(CHEST_LEVELS, row):
        if cell.value:
            private_config['coded_chest'][lvl].append(cell.value)

ws = wb['BlitzCoin']
coins = game_config['coins'] = {}

coin_time = coins['time'] = []
for row in ws['B2': 'C15']:
    coin_time.append({
            'time': row[0].value,
            'coin': row[1].value
        })

coin_resource = coins['resource'] = []
for row in ws['E2': 'F11']:
    coin_resource.append({
            'resource': row[0].value,
            'coin': row[1].value
        })


ws = wb['Modules']
modules = game_config['modules'] = {}
cur_module = None
for row in ws['A1':'C200']:
    if row[0].value:
        cur_module = {
            'type': 'unit'
        }
        modules[row[0].value] = cur_module

    if row[1].value:
        cur_module[row[1].value] = row[2].value

cur_module = None
for row in ws['E1':'G200']:
    if row[0].value:
        cur_module = {
            'type': 'tower'
        }
        modules[row[0].value] = cur_module

    if row[1].value:
        cur_module[row[1].value] = row[2].value

ws = wb['Radar']
game_config['radar'] = {}

for pos, row in enumerate(collect_table(ws['A1': 'O1'], ws['A2':'O100'])):
    name = row.pop('name')
    row.pop('distance')
    row['order'] = pos
    
    rewards = {
        'resources': {
            'adamantite': row.pop('rw.ad', None),
            'crystalite': row.pop('rw.cr', None),
            'titanium': row.pop('rw.ti', None),
        }
    }
    chest_lvl = row.pop('rw.ch', None)
    if chest_lvl:
        rewards['chest'] = 'lvl{}'.format(chest_lvl)

    if row['type'] == 'base':
        row['rewards'] = rewards

    game_config['radar'][name] = row
    

ws = wb['Expedition']
game_config['expedition'] = []

for pos, row in enumerate(collect_table(ws['A1': 'I1'], ws['A2':'I100'])):
    row['codingPos'] = int(100*row['codingPos'])
    row['chlvl1'] = int(100*row['chlvl1'])
    row['chlvl2'] = int(100*row['chlvl2'])
    row['chlvl3'] = int(100*row['chlvl3'])

    game_config['expedition'].append(row)

if os.path.exists(BALANCE_JSON):
    with open(BALANCE_JSON, 'w') as fh:
        json.dump(game_config, fh, indent=2)

if os.path.exists(PRIVATE_BALANCE_JSON):
    with open(PRIVATE_BALANCE_JSON, 'w') as fh:
        json.dump(private_config, fh, indent=2)

with open(LOCAL_BALANCE_JSON, 'w') as fh:
    json.dump(game_config, fh, indent=2)

with open(LOCAL_PRIVATE_BALANCE_JSON, 'w') as fh:
    json.dump(private_config, fh, indent=2)


if os.path.exists(BALANCE_BASES_FOLDER):
    strategies = {}
    for filename in os.listdir(STRATEGIES_FOLDER):
        with open(os.path.join(STRATEGIES_FOLDER, filename)) as fh:
            strategies[filename] = {
                'at': 0,
                'content': fh.read()
            }
    
    for filename in os.listdir(BASES_FOLDER):
        with open(os.path.join(BASES_FOLDER, filename)) as fh:
            base_data = json.load(fh)
            base_data['strategies'] = strategies
            base_data['profile']['username'] = filename.split('.')[0]
        
        with open(os.path.join(BALANCE_BASES_FOLDER, filename), 'w') as fh:
            json.dump(base_data, fh)